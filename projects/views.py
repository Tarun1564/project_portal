from django.http import HttpResponseForbidden, JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from .models import *
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from accounts.models import CustomUser
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.db.models import Avg
from django.db import IntegrityError
import pandas as pd
import uuid
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
@login_required
def coordinator_dashboard(request):
    graduation=Graduation.objects.all()
    branches=Branch.objects.all()
    academic_years=AcademicYear.objects.all()
    context = {
        'graduation':graduation,
        'branches':branches,
        'academic_years':academic_years,
    }
    return render(request, 'dashboard_coordinator.html', context)
@login_required
def student_dashboard(request):
    graduation=Graduation.objects.all()
    branches=Branch.objects.all()
    academic_years=AcademicYear.objects.all()
    context = {
        'graduation':graduation,
        'branches':branches,
        'academic_years':academic_years,
    }
    return render(request, 'dashboard_student.html', context)

@login_required
def upload_batches_excel(request):
    if request.method != "POST" or "excel_file" not in request.FILES:
        messages.error(request, "Please upload a valid Excel file.")
        return redirect("coordinator_dashboard")

    try:
        df = pd.read_excel(request.FILES["excel_file"])
    except Exception:
        messages.error(request, "Invalid Excel file.")
        return redirect("coordinator_dashboard")

    df.columns = df.columns.str.strip()
    required_columns = [
        "Batch No",
        "Graduation",
        "Branch",
        "Academic Year",
        "Name of the Guide",
        "Name of the Coordinator",
        "Roll No",
        "Name of the Student",
    ]

    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        messages.error(request, f"Missing columns: {', '.join(missing)}")
        return redirect("coordinator_dashboard")
    df["Graduation"] = df["Graduation"].ffill()
    df["Branch"] = df["Branch"].ffill()
    df["Academic Year"] = df["Academic Year"].ffill()
    df["Batch No"] = df["Batch No"].ffill()
    df["Name of the Guide"] = df["Name of the Guide"].ffill()
    df["Name of the Coordinator"] = df["Name of the Coordinator"].ffill()

    if "Course" in df.columns:
        df["Course"] = df["Course"].ffill()

    created_batches = 0
    created_students = 0
    with transaction.atomic():

        for _, row in df.iterrows():
            graduation, _ = Graduation.objects.get_or_create(
                name=str(row["Graduation"]).strip()
            )
            course = None
            course_name = ""

            if "Course" in df.columns:
                course_name = str(row.get("Course", "")).strip()

            if graduation.name.lower() == "diploma":
                course = None
            else:
                if not course_name:
                    raise ValueError(
                        f"Course is mandatory for {graduation.name}"
                    )

                course, _ = Course.objects.get_or_create(
                    name=course_name,
                    graduation=graduation
                )
            branch, _ = Branch.objects.get_or_create(
                graduation=graduation,
                course=course,
                name=str(row["Branch"]).strip()
            )
            academic_year, _ = AcademicYear.objects.get_or_create(
                year=str(row["Academic Year"]).strip()
            )
            guide_name = str(row["Name of the Guide"]).strip()
            coordinator_name = str(row["Name of the Coordinator"]).strip()

            coordinator_user, created = CustomUser.objects.get_or_create(
                username=coordinator_name.replace(" ", "").lower(),
                defaults={"user_role": "coordinator"},
            )

            if created:
                coordinator_user.set_password("defaultpassword123")
                coordinator_user.save()

            coordinator, _ = Coordinator.objects.get_or_create(
                user=coordinator_user,
                defaults={
                    "name": coordinator_name,
                    "department": branch
                }
            )
            batch, batch_created = Batch.objects.get_or_create(
                number=row["Batch No"],
                graduation=graduation,
                course=course,
                branch=branch,
                academic_year=academic_year,
                defaults={
                    "guide": guide_name,
                    "coordinator": coordinator
                }
            )

            if not batch_created:
                batch.guide = guide_name
                batch.coordinator = coordinator
                batch.save()
            else:
                created_batches += 1
            roll_no = str(row["Roll No"]).strip()

            student, student_created = Student.objects.get_or_create(
                roll_number=roll_no,
                defaults={
                    "name": str(row["Name of the Student"]).strip(),
                    "graduation": graduation,
                    "course": course,
                    "branch": branch,
                    "academic_year": academic_year,
                    "guide": guide_name,
                    "coordinator": coordinator,
                    "batch": batch,
                },
            )

            if not student_created:
                student.batch = batch
                student.save()
            else:
                created_students += 1

            batch.members.add(student)

    messages.success(
        request,
        f"Upload successful: {created_batches} batches and {created_students} students added."
    )

    return redirect("coordinator_dashboard")

@login_required
def approve_project(request, project_id):
    coordinator = get_object_or_404(Coordinator, user=request.user)

    project = get_object_or_404(
        Project,
        id=project_id,
        batch__coordinator=coordinator, 
        coordinator_approved=False
    )

    project.coordinator_approved = True
    project.save(update_fields=['coordinator_approved'])

    return redirect('coordinator_dashboard')
@login_required
def reject_project(request, project_id):
    coordinator = get_object_or_404(Coordinator, user=request.user)
    project = get_object_or_404(
        Project,
        id=project_id,
        batch__coordinator=coordinator
    )
    batch = project.batch
    batch.is_submitted = False
    batch.save(update_fields=['is_submitted'])
    if project.abstract:
        project.abstract.delete(save=False)
    if project.documentation:
        project.documentation.delete(save=False)
    if project.presentation:
        project.presentation.delete(save=False)
    project.delete()
    return redirect('coordinator_dashboard')

@login_required
def update_project_rating(request, project_id):
    if request.method == "POST":
        project = get_object_or_404(Project, id=project_id)
        if request.user.user_role != "coordinator":
            return redirect(request.META.get("HTTP_REFERER"))

        rating = request.POST.get("rating")

        if rating:
            project.rating = abs(int(rating))
            project.save(update_fields=["rating"])

        return redirect(request.META.get("HTTP_REFERER"))
@login_required
def get_courses(request, graduation_id):
    courses = Course.objects.filter(graduation_id=graduation_id)
    return JsonResponse(list(courses.values("id","name")), safe=False)

@login_required
def get_branches(request, graduation_id, course_id):
    if course_id == 0:
        branches = Branch.objects.filter(
            graduation_id=graduation_id,
            course__isnull=True
        )
    else:
        branches = Branch.objects.filter(course_id=course_id)

    return JsonResponse(
        list(branches.values("id", "name")),
        safe=False
    )

@login_required
def get_academic_years(request, branch_id):
    academic_years = AcademicYear.objects.all().order_by('-year')

    data = [{"id": y.id, "year": y.year} for y in academic_years]
    return JsonResponse(data, safe=False)
@login_required
def get_batches(request, year_id, branch_id):

    if request.user.user_role == "coordinator":
        batches = Batch.objects.filter(
            academic_year_id=year_id,
            branch_id=branch_id,
            coordinator__user=request.user
        )
    else:
        batches = Batch.objects.filter(
            academic_year_id=year_id,
            branch_id=branch_id
        )

    batches = batches.select_related(
        'coordinator__user',
        'project'
    ).prefetch_related(
        'members'
    )

    data = []
    for batch in batches:
        project = getattr(batch, "project", None)
        project_entries = None
        if project:
            project_entries = {
                "id": project.id,
                "title": project.title,
                "rating": project.rating,
                "coordinator_approved": project.coordinator_approved,
                "abstract": project.abstract.url if project.abstract else "",
                "documentation": project.documentation.url if project.documentation else "",
                "presentation": project.presentation.url if project.presentation else "",
                "nptel": project.nptel.url if project.nptel else "",
                "paper": project.paper.url if project.paper else "",
                "status": "Approved" if project.coordinator_approved else "Pending"
            }
        students = [{
            "name": s.name,
            "roll_number": s.roll_number,
            "guide": s.guide,
            "coordinator": (
                s.coordinator.user.username
                if s.coordinator else "-"
            )
        } for s in batch.members.all()]
        data.append({
            "batch_id": batch.id,
            "batch_number": batch.number,
            "guide": batch.guide,
            "members_count": batch.members.count(),
            "students": students,   # ✅ ADDED
            "coordinator": (
                batch.coordinator.user.username
                if batch.coordinator else "-"
            ),
            "project_entries": project_entries
        })

    return JsonResponse(data, safe=False)

@login_required
@csrf_exempt
def edit_project_files(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == "POST":
        updated = False
        if 'abstract' in request.FILES:
            project.abstract = request.FILES['abstract']
            updated = True
        if 'documentation' in request.FILES:
            project.documentation = request.FILES['documentation']
            updated = True
        if 'ppt' in request.FILES:
            project.presentation = request.FILES['ppt']
            updated = True
        if 'nptel' in request.FILES:
            project.nptel=request.FILES['nptel']
            updated=True
        if 'research_paper' in request.FILES:
            project.paper=request.FILES['research_paper']
            updated=True
        if updated:
            project.save()
            return JsonResponse({"status":"success"})
        else:
            return JsonResponse({"status":"no_files_uploaded"}, status=400)

    return JsonResponse({"status":"invalid_request"}, status=400)
@login_required
def batch_students(request, batch_id):
    batch = get_object_or_404(Batch, id=batch_id)
    students = batch.students.all()

    data = []
    for s in students:
        data.append({
            "name": s.name,
            "roll_number": s.roll_number,
            "guide": s.guide,
            "coordinator": s.coordinator.user.username if s.coordinator else "-"
        })
    return JsonResponse(data, safe=False)
@login_required
def download_data(request, year_id, branch_id):
    batches = (
        Batch.objects
        .filter(academic_year_id=year_id, branch_id=branch_id)
        .prefetch_related('students', 'batch_projects')
    )

    academic_year = AcademicYear.objects.filter(id=year_id).first()
    branch = Branch.objects.filter(id=branch_id).first()
    return "Done"
@login_required
def upload_project_batch(request, batch_id):

    batch = get_object_or_404(Batch, id=batch_id)

    if request.method != "POST":
        return JsonResponse({"status": "error", "error": "Invalid request"})
    if batch.is_submitted==True:
        return JsonResponse({"status":"error","error":"Project Already Submitted"})
    title = request.POST.get("title")
    abstract = request.FILES.get("abstract")
    documentation = request.FILES.get("documentation")
    presentation = request.FILES.get("presentation")
    nptel_certifications = request.FILES.get("nptel_certifications")
    paper = request.FILES.get("research_paper")

    if not title or not abstract or not documentation or not presentation:
        return JsonResponse({
            "status": "error",
            "error": "All required fields must be uploaded."
        })

    project = Project.objects.create(
        batch=batch,
        title=title,
        abstract=abstract,
        documentation=documentation,
        presentation=presentation,
        nptel=nptel_certifications,
        paper=paper  
    )

    batch.is_submitted = True
    batch.save(update_fields=["is_submitted"])

    return JsonResponse({"status": "success"})
def guide_feedback(request, year_id, feedback_uuid):
    if Branch.objects.filter(feedback_uuid=feedback_uuid).exists():
        branch = Branch.objects.get(feedback_uuid=feedback_uuid)
        batches=Batch.objects.filter(branch=branch,academic_year_id=year_id)
        return render(request, 'feedback.html', {
            'batches': batches
        })
    return HttpResponseForbidden("Invalid or expired feedback link.")
def get_project_paper(request):
    batch_id = request.GET.get("batch_id")

    try:
        project = Project.objects.get(batch_id=batch_id)
        return JsonResponse({
            "success": True,
            "paper_url": project.paper.url,
            "project_id": project.id
        })
    except Project.DoesNotExist:
        return JsonResponse({"success": False})
def upload_feedback(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method != "POST":
        return HttpResponseForbidden("Invalid request")
    name = request.POST.get("name", "")
    evaluator_id = request.POST.get("email", "")
    innovation = int(request.POST.get("innovation", 0))
    mixed_methods = int(request.POST.get("mixed", 0))
    society_impact = int(request.POST.get("society_impact", 0))
    feasibility = int(request.POST.get("feasibility", 0))
    outcome_harvesting = int(request.POST.get("outcome_harvesting", 0))
    paper_based_selection = int(request.POST.get("paper_based_selection", 0))
    sustainability = int(request.POST.get("sustainability", 0))
    weights = {
        "innovation": 25,
        "mixed": 25,
        "society_impact": 15,
        "feasibility": 10,
        "outcome_harvesting": 10,
        "paper_based_selection": 10,
        "sustainability": 5,
    }
    total_score = (
        innovation * weights["innovation"]
        + mixed_methods * weights["mixed"]
        + society_impact * weights["society_impact"]
        + feasibility * weights["feasibility"]
        + outcome_harvesting * weights["outcome_harvesting"]
        + paper_based_selection * weights["paper_based_selection"]
        + sustainability * weights["sustainability"]
    ) / 10

    try:
        Feedback.objects.create(
            project=project,
            evaluator_name=name,
            evaluator_id=evaluator_id,
            rating=round(total_score, 2)
        )
    except IntegrityError:
        return HttpResponse("<h3 style='text-align:center;margin-top:40px;'>⚠️ You already submitted feedback.</h3>")
    avg_rating = project.feedbacks.aggregate(Avg("rating"))["rating__avg"]
    project.rating = round(avg_rating, 2)
    project.save(update_fields=["rating"])
    return HttpResponse("<h3 style='text-align:center;margin-top:40px;'>✅ Thank you for your feedback!</h3>")
def generate_feedback_link(request, year_id, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    year = get_object_or_404(AcademicYear, id=year_id)
    link = request.build_absolute_uri(
        reverse('guide_feedback', args=[year_id, branch.feedback_uuid])
    )
    return JsonResponse({
        "status": "success",
        "feedback_link": link
    })
def get_feedback_details(request,project_id):
    project = get_object_or_404(Project, id=project_id)
    feedbacks = project.feedbacks.all()
    data = []
    for fb in feedbacks:
        data.append({
            "evaluator_name": fb.evaluator_name,
            "evaluator_id": fb.evaluator_id,
            "rating": fb.rating,
        })
    return JsonResponse({"feedbacks": data})
def download_batch_excel(request, year_id, branch_id):
    branch = Branch.objects.filter(id=branch_id).first()
    academic_year = AcademicYear.objects.filter(id=year_id).first()
    top_projects = Project.objects.filter(
    batch__academic_year_id=year_id,
    batch__branch_id=branch_id
).order_by("-rating")[:5]

    top_ids = [p.id for p in top_projects]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Batch Projects"

    base_url = request.build_absolute_uri("/")[:-1]

    def set_link(sheet, row, col, text, url):
        if url:
            cell = sheet.cell(row=row, column=col)
            cell.value = text
            cell.hyperlink = url
            cell.style = "Hyperlink"
        else:
            sheet.cell(row=row, column=col, value="-")

    sheet["A1"] = "ADITYA INSTITUTE OF TECHNOLOGY AND MANAGEMENT (A), TEKKALI - 532201"
    sheet["A2"] = branch.name
    sheet["A3"] = academic_year.year
    sheet["A4"] = "PROJECT BATCHES"

    for h in ["A1", "A2", "A3", "A4"]:
        sheet[h].font = Font(size=13, bold=True)
        sheet.merge_cells(f"{h}:L{h[1]}")
        sheet[h].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "SNO", "Batch Number", "Roll Number", "Name Of The Student",
        "Name Of The Guide", "Title Of The Project", "Abstract",
        "Documentation", "Research Paper", "Presentation",
        "NPTEL Certifications", "Rating","Name Of The Coordinator"
    ]
    for col, title in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=col)
        cell.value = title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    batches = Batch.objects.filter(
        academic_year=academic_year,
        branch=branch
    ).select_related("project").prefetch_related("members")

    cur_row = 6
    sno = 1
    
    for batch in batches:
        students = batch.members.all()
        count = students.count()
        start_row = cur_row
        project = getattr(batch, "project", None)
        sheet.cell(row=start_row, column=1, value=sno)
        sheet.cell(row=start_row, column=2, value=batch.number)
        sheet.cell(row=start_row,column=2).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=start_row, column=5, value=batch.guide)
        if project:
            sheet.cell(row=start_row, column=6, value=project.title)

            set_link(sheet, start_row, 7, "View Abstract",
                     base_url + project.abstract.url if project.abstract else "")

            set_link(sheet, start_row, 8, "View Docs",
                     base_url + project.documentation.url if project.documentation else "")

            set_link(sheet, start_row, 9, "View Paper",
                     base_url + project.paper.url if project.paper else "")

            set_link(sheet, start_row, 10, "View PPT",
                     base_url + project.presentation.url if project.presentation else "")

            set_link(sheet, start_row, 11, "View NPTEL",
                     base_url + project.nptel.url if project.nptel else "")

            sheet.cell(row=start_row, column=12, value=project.rating)
        else:
            sheet.cell(row=start_row, column=6, value="Not Submitted")

        for student in students:
            sheet.cell(row=cur_row,column=1,value=sno)
            sheet.cell(row=cur_row, column=3, value=student.roll_number)
            sheet.cell(row=cur_row, column=4, value=student.name)
            sno+=1
            cur_row += 1

        if count > 1:
            for col in [ 2, 5, 6, 7, 8, 9, 10, 11, 12]:
                sheet.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=start_row + count - 1,
                    end_column=col
                )
            yellow_fill = PatternFill(
            start_color="FFF9C4",
            end_color="FFF9C4",
            fill_type="solid"
                )

        if project and project.id in top_ids:
            for r in range(start_row, start_row + count):
                for c in range(1, 13):
                    sheet.cell(row=r, column=c).fill = yellow_fill

        sno += 1
        column_widths = {
    "A": 6,
    "B": 14,
    "C": 18,
    "D": 28,
    "E": 24,
    "F": 35,
    "G": 18,
    "H": 20,
    "I": 22,
    "J": 18,
    "K": 22,
    "L": 12,
    "M":20
}
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
    sheet.cell(row=6,column=13,value=request.user.username)
    sheet.merge_cells(start_row=6,start_column=13,end_row=6+sno-1,end_column=13)
    sheet.cell(row=6,column=13).alignment=Alignment(horizontal="center",vertical="center")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="Batch_Submissions_{academic_year.year}.xlsx"'
    )

    workbook.save(response)
    return response
import zipfile
import os
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .models import Project, Coordinator


@login_required
def download_all_nptel(request):
    # coordinator profile
    coordinator = Coordinator.objects.get(user=request.user)

    branch = coordinator.department
    year = None 

    projects = Project.objects.filter(
        batch__branch=branch
    ).exclude(nptel="")

    if not projects.exists():
        return HttpResponse("No NPTEL certificates found")

    filename = f"{branch.name}_NPTEL.zip"
    response = HttpResponse(content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    zip_file = zipfile.ZipFile(response, 'w', zipfile.ZIP_DEFLATED)

    for project in projects:
        if project.nptel and os.path.exists(project.nptel.path):
            batch_no = project.batch.number
            year = project.batch.academic_year.year

            arcname = f"{year}/Batch_{batch_no}.pdf"
            zip_file.write(project.nptel.path, arcname)

    zip_file.close()
    return response